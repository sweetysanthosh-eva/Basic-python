import openpyxl


def readdata():
    book=openpyxl.load_workbook("C:\\Users\\sweet\\Downloads\\TestData.xlsx")
    sheet=book.active
    if"LoginPage" in book.sheetnames:
        sheet=book.worksheets[0]
        username=sheet.cell(row=1,column=1).value
    return username
print(readdata())